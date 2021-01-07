import pyzbar.pyzbar as pyzbar
import cv2
import matplotlib.pyplot as plt

from openpyxl import Workbook
from datetime import datetime, timezone, date, time
from pytz import timezone, utc

KST = timezone('Asia/Seoul')
split_day_time = []
#now_time = datetime.utcnow()
list_h_m_s = []
#y_m_d = utc.localize(now_time).astimezone(KST).strftime("%Y-%m-%d")
#print(y_m_d)

#print(now_time.strftime("%Y-%m-%d"))
'''h_m_s = utc.localize(now_time).astimezone(KST).strftime("%H-%M-%d")'''
#h_m_s = utc.localize(now_time).astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")

#list_h_m_s.append(datetime.strptime(h_m_s,"%Y-%m-%d %H:%M:%S"))

time_a  = '2021-01-05 06:00:00'
name_list = ['Yeon', 'Chan', 'Min']
#print(list_h_m_s[0])
#list_h_m_s.append(datetime.strptime('2021-01-05 07:00:00',"%Y-%m-%d %H:%M:%S"))

a_time = datetime.strptime(time_a,"%Y-%m-%d %H:%M:%S")

img = cv2.imread('/home/hwangbo1998/qr_y.png')

gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

decoded = pyzbar.decode(gray)

print(decoded)

for d in decoded :
	personal_data = d.data.decode('utf-8')
	decoded_type = d.type

while 1 :
	x = input()
	if x == 'now' :
		if decoded_type and personal_data in name_list :
			
			print(personal_data)
			now_time = datetime.utcnow()
			fake_h_m_s = utc.localize(now_time).astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")
			list_h_m_s.append(datetime.strptime(fake_h_m_s,"%Y-%m-%d %H:%M:%S"))
			print(list_h_m_s)
	elif x == 'bye' :
		break
	else :
		pass
# just add the time easily. i will delete the while line.

#diff = list_h_m_s[0] - a_time
#print(diff.total_seconds()/60)

for i in range(1,len(list_h_m_s)) :
	abs_diff_1 = abs( list_h_m_s[0] - list_h_m_s[i])
	diff_sec = abs_diff_1.total_seconds()/60
	if diff_sec >= 1 : #test execute >1 eigentlich <1 ist richtig!
		split_day_time = list_h_m_s[0].strftime("%Y-%m-%d %H:%M:%S").split()
		print(split_day_time[0])
		print(split_day_time[1])		
		break


#print(split_day_time[0])
#print(split_day_time[1])
print(len(list_h_m_s))
print(list_h_m_s[0])
print(list_h_m_s[1]) 
print(diff_sec)

#print(h_m_s)

print(now_time.strftime("%H:%M:%S"))


'''send_excel()

def send_excel() :
        write_wb = Workbook()
        write_ws = write_wb.create_sheet('Jan,Feb')

        print(wb.sheetnames)

        write_ws = write_wb.active
        now_time = datetime.now()

	today.isoformat()

'''
