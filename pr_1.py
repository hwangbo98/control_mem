import pyzbar.pyzbar as pyzbar
import cv2
import matplotlib.pyplot as plt

from openpyxl import Workbook
from datetime import datetime

img = cv2.imread('/home/hwangbo1998/qr_y.png')

plt.imshow(img)

gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

plt.imshow(img,cmap='gray')

decoded = pyzbar.decode(gray)

print(decoded)

for d in decoded :
	print(d.data.decode('utf-8'))
	print(d.type)

	cv2.rectangle(img,(d.rect[0],d.rect[1]),(d.rect[0] + d.rect[2], d.rect[1] + d.rect[3]), (0,0,255),2)

plt.imshow(img)


def send_excel() :
	write_wb = Workbook()
	write_ws = write_wb.create_sheet('Jan,Feb')

	print(wb.sheetnames)

	write_ws = write_wb.active
	
	now_time = datetime.now()

	
	
	


